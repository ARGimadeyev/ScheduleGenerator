import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook('data/112.xlsx', data_only=True)

sheet = wb.active
classes = dict()
subjects = dict()
s = dict()
numSub = 1
teachers = dict()
t = dict()
numTeach = 1
for i in range(3, sheet.max_column):
    if not sheet.cell(row=1, column=i): continue
    classes[i] = sheet.cell(row=1, column=i).value.replace('\n', '').replace('/', '.').strip()
result = []
lastTeacher = None
for i in range(2, sheet.max_row + 1):
    subject = sheet.cell(row=i, column=1).value
    if not subject: continue
    subject = str(subject).strip().capitalize()
    teacher = sheet.cell(row=i, column=2).value
    if not teacher: teacher = lastTeacher
    teacher = str(teacher).strip()
    if subject not in subjects:
        subjects[subject] = numSub
        if (numSub == 3): subjects[subject] = 2
        s[numSub] = subject
        numSub += 1
    if teacher not in teachers:
        teachers[teacher] = numTeach
        t[numTeach] = teacher
        numTeach += 1
    for j in range(3, sheet.max_column):
        if not sheet.cell(row=i, column=j).value: continue
        col = sheet.cell(row=i, column=j).value
        dop = sheet.cell(row=i, column=j).fill.fgColor.rgb != '00FFFF00'
        if dop:
            dop = 1
        else:
            dop = 0
        for h in range(int(col)):
            result.append([teachers[teacher], subjects[subject], classes[j], dop])
    lastTeacher = teacher
with open('data/ListLesson.txt', 'w', encoding='utf-8') as f:
    for row in result:
        for item in row:
            f.write(str(item) + ' ')
        f.write('\n')
with open('data/maps/teachers.json', 'w', encoding='utf-8') as f:
    json.dump(t, f, ensure_ascii=False, indent=2)

with open('data/maps/subjects.json', 'w', encoding='utf-8') as f:
    json.dump(s, f, ensure_ascii=False, indent=2)
